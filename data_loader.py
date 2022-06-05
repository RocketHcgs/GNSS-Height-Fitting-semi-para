from openpyxl import load_workbook
import numpy as np

data_file = 'data.xlsx' #原始数据文件
output_file = 'data.xlsx' #输出数据文件
# 拟合点数据位置设置
posX_min = 'B2'
posX_max = 'B16'
posY_min = 'C2'
posY_max = 'C16'
posZeta_min = 'F2'
posZeta_max = 'F16'
# 验证数据位置设置
posX_min_v = 'B17'
posX_max_v = 'B25'
posY_min_v = 'C17'
posY_max_v = 'C25'
posZeta_min_v = 'F17'
posZeta_max_v = 'F25'
# 输出位置
out_row = 1
out_col = 9

def load_data():
	wb = load_workbook(data_file)
	ws = wb.active
	cellX, cellY, cellZeta = [], [], []

	for i in ws[posX_min:posX_max]:
		cellX.append(i[0].value)
	for i in ws[posY_min:posY_max]:
		cellY.append(i[0].value)
	for i in ws[posZeta_min:posZeta_max]:
		cellZeta.append(i[0].value)

	matX = np.mat(cellX)
	matY = np.mat(cellY)
	matZeta = np.mat(cellZeta)
	return (matX.T, matY.T, matZeta.T) #返回值分别为x,y,ζ组成的列向量

def load_verify_data():
	wb = load_workbook(data_file)
	ws = wb.active
	cellX, cellY, cellZeta = [], [], []

	for i in ws[posX_min_v:posX_max_v]:
		cellX.append(i[0].value)
	for i in ws[posY_min_v:posY_max_v]:
		cellY.append(i[0].value)
	for i in ws[posZeta_min_v:posZeta_max_v]:
		cellZeta.append(i[0].value)

	matX_v = np.mat(cellX)
	matY_v = np.mat(cellY)
	matZeta_v = np.mat(cellZeta)
	return (matX_v.T, matY_v.T, matZeta_v.T)

def save_data(matA, matB, matC):
	wb = load_workbook(output_file)
	ws = wb.active
	ws.cell(row = out_row, column = out_col, value = "out1")
	ws.cell(row = out_row + matA.shape[0] + 1, column = out_col, value = "out2")
	ws.cell(row = out_row + matA.shape[0] + matB.shape[0] + 2, column = out_col, value = "out3")

	for i in range(matA.shape[0]):
		for j in range(matA.shape[1]):
			ws.cell(row = out_row + 1 + i, column = out_col + j, value = matA[i,j])
	for i in range(matB.shape[0]):
		for j in range(matB.shape[1]):
			ws.cell(row = out_row + matA.shape[0] + 2 + i, column = out_col + j, value = matB[i,j])
	for i in range(matC.shape[0]):
		for j in range(matC.shape[1]):
			ws.cell(row = out_row + matA.shape[0] + matB.shape[0] + 3 + i, column = out_col + j, value = matC[i,j])
	wb.save(output_file)

if __name__ == "__main__":
	x, y, zeta = load_data()
	print("number of fitting data:", x.shape[0])
	x, y, zeta = load_verify_data()
	print("number of verify data:", x.shape[0])